import os
import random
import json
import re
import hashlib
from openpyxl import Workbook, load_workbook

# ======================== CONFIGURATION ========================
ROOT_DIRECTORY = 'Content'  #Location of your folder if not in the same directory

# ======================== UTILITY FUNCTIONS ========================
def validate_inscription_id(inscription_id):
    return re.match(r"^[\da-fA-F]{64}i\d+$", inscription_id) is not None

def get_positive_integer(message):
    while True:
        try:
            value = int(input(message))
            if value > 0:
                return value
            else:
                print("Please enter a positive number.")
        except ValueError:
            print("Invalid input. Please enter a valid number.")

# ======================== SPREADSHEET MANAGEMENT ========================
def create_spreadsheet(traits_dict):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Traits Information"

    headers = ["Trait Number", "Trait Type", "Trait", "Inscription ID", "Rarity (%)", "Avoid Traits (use Trait Numbers, comma-separated)", "Use None (Yes/No)"]
    sheet.append(headers)

    trait_number = 1
    for trait_type, traits in traits_dict.items():
        for trait in traits:
            row = [trait_number, trait_type, trait, '', '', '', '']
            sheet.append(row)
            trait_number += 1
        sheet.append([trait_number, trait_type, 'none', '', '', '', ''])
        trait_number += 1

    workbook.save("traits_info.xlsx")
    print("\nA file named 'traits_info.xlsx' has been created. Please fill in the required information in this file before proceeding.")

def load_traits_info():
    workbook = load_workbook("traits_info.xlsx")
    sheet = workbook.active

    all_traits_info = {}
    trait_number_mapping = {}
    errors = []
    cumulative_rarity = {}

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue

        trait_number, trait_type, trait, inscription_id, rarity, avoid_traits, use_none = (item if item is not None else '' for item in row)

        if inscription_id and not validate_inscription_id(inscription_id):
            errors.append(f"Row {row_index}: Invalid Inscription ID format.")
            continue

        if rarity != '':
            try:
                rarity = float(rarity) 
                if not (0 <= rarity <= 100):
                    errors.append(f"Row {row_index}: Rarity must be a value between 0 and 100.")
                    continue
            except ValueError:
                errors.append(f"Row {row_index}: Rarity should be a number or left blank for 'Random'.")
                continue
        else:
            rarity = 'Random'
        if isinstance(rarity, float):
            cumulative_rarity[trait_type] = cumulative_rarity.get(trait_type, 0) + rarity

        avoid_list = []
        if avoid_traits:
            avoid_traits_str = str(avoid_traits)
            avoid_traits_list = avoid_traits_str.split(',') if ',' in avoid_traits_str else [avoid_traits_str]

            for potential_number in avoid_traits_list:
                stripped_value = potential_number.strip()
                try:
                    number_value = float(stripped_value)
                    if number_value.is_integer():
                        avoid_list.append(int(number_value))
                    else:
                        errors.append(f"Row {row_index}: Avoid Traits contains invalid entry '{stripped_value}'. Only whole numbers are allowed.")
                        break
                except ValueError:
                    errors.append(f"Row {row_index}: Avoid Traits contains invalid entry '{stripped_value}'. Only numbers are allowed.")
                    break

        if errors:
            continue

        if use_none.strip().lower() not in ('', 'yes', 'no'):
            errors.append(f"Row {row_index}: 'Use None' must be either 'Yes', 'No', or left blank.")
            continue

        use_none_flag = use_none.strip().lower() == 'yes'

        if trait_type not in all_traits_info:
            all_traits_info[trait_type] = {}

        all_traits_info[trait_type][trait] = {
            "inscription_id": inscription_id,
            "rarity": rarity,
            "avoid_traits": avoid_list,
            "use_none": use_none_flag
        }

        trait_number_mapping[int(trait_number)] = (trait_type, trait)
    for trait_type, total_rarity in cumulative_rarity.items():
        if not (99.9 <= total_rarity <= 100.1): 
            errors.append(f"Total rarity for '{trait_type}' is not 100%. It is currently {total_rarity}%.")

    if errors:
        print("\nErrors were found in the spreadsheet. Please review the messages below.")
        for error_message in errors:
            print(error_message)
        print("\nPlease correct the errors in the spreadsheet and try again.")
        exit()

    return all_traits_info, trait_number_mapping

# ======================== VALIDATION FUNCTIONS ========================
def validate_inscription_avoidance(inscription_collection, all_traits_info, trait_number_mapping):
    inconsistencies = []

    for inscription_index, inscription in enumerate(inscription_collection, start=1):
        current_inscription_trait_numbers = set()

        for trait in inscription:
            trait_type = trait["trait_type"]
            trait_value = trait["value"]
            trait_number = next((number for number, (t_type, t_name) in trait_number_mapping.items() if t_type == trait_type and t_name == trait_value), None)

            if trait_number is not None:
                current_inscription_trait_numbers.add(trait_number)

        for trait in inscription:
            trait_type = trait["trait_type"]
            trait_value = trait["value"]
            trait_details = all_traits_info[trait_type][trait_value]

            avoid_traits = set(trait_details.get("avoid_traits", []))
            conflicting_traits = current_inscription_trait_numbers & avoid_traits

            if conflicting_traits:
                conflicts = [f"Trait #{num}" for num in conflicting_traits]
                inconsistency_info = {
                    "Inscription_number": inscription_index,
                    "Trait": f"{trait_type} - {trait_value}",
                    "Conflicts": conflicts
                }
                inconsistencies.append(inconsistency_info)

    return inconsistencies

def validate_traits(selected_traits, all_traits_info, trait_number_mapping):
    for trait_type, trait_name in selected_traits.items():
        current_trait_info = all_traits_info[trait_type][trait_name]
        avoid_list = current_trait_info.get("avoid_traits", [])

        for avoid_trait_number in avoid_list:
            avoid_trait_type, avoid_trait_name = trait_number_mapping[avoid_trait_number]

            if selected_traits.get(avoid_trait_type) == avoid_trait_name:
                return False

    return True

# ======================== METADATA GENERATION ========================

def generate_inscriptions(all_traits_info, trait_number_mapping, num_inscriptions):
    inscription_collection = []
    traits_usage = {trait_type: {trait: {"count": 0, "rarity": traits_info[trait]["rarity"]} for trait in traits_info} for trait_type, traits_info in all_traits_info.items()}

    generated_hashes = set()
    trait_count_distribution = {}

    for _ in range(num_inscriptions):
        generation_attempts = 0

        while True:
            inscription_traits = {}
            formatted_traits = []
            current_avoid_list = []

            for trait_type, traits in all_traits_info.items():
                available_traits = {trait_name: trait_details for trait_name, trait_details in traits.items() if trait_number_mapping.get((trait_type, trait_name)) not in current_avoid_list}

                if not available_traits:
                    raise Exception(f"Conflict in 'Avoid Traits' rules prevents generation of valid metadata. Please revise the rules.")

                probabilities = []
                trait_names = []

                for trait_name, trait_info in available_traits.items():
                    if trait_name == 'none' and not trait_info['use_none']:
                        continue

                    rarity = trait_info.get("rarity")

                    if rarity == 'Random' or rarity == '':
                        probabilities.append(1.0)
                    else:
                        try:
                            rarity_value = float(rarity)
                            probabilities.append(rarity_value)
                        except ValueError:
                            print(f"Invalid rarity value detected for trait {trait_name}. It should be a number.")
                            return

                    trait_names.append(trait_name)

                selected_trait = random.choices(trait_names, weights=probabilities, k=1)[0]
                selected_trait_info = traits[selected_trait]

                current_avoid_list.extend(selected_trait_info.get("avoid_traits", []))

                if selected_trait != 'none':
                    inscription_traits[trait_type] = selected_trait
                    formatted_traits.append({
                        "trait_type": trait_type,
                        "value": selected_trait
                    })

            inscription_hash = hashlib.sha256(str(inscription_traits).encode()).hexdigest()
            if inscription_hash in generated_hashes:
                generation_attempts += 1
                print("Duplicate metadata detected. Reselecting with different traits...")

                if generation_attempts >= 10000:
                    print(f"Unable to generate unique metadata after {generation_attempts} attempts. You may have exhausted all unique combinations. Please add more traits, or lower collection total size.")
                    return inscription_collection, traits_usage, trait_count_distribution

                continue

            generated_hashes.add(inscription_hash)

            valid_inscription = validate_traits(inscription_traits, all_traits_info, trait_number_mapping)

            if valid_inscription:
                for trait_type, selected_trait in inscription_traits.items():
                    traits_usage[trait_type][selected_trait]["count"] += 1

                for trait_type in all_traits_info:
                    if trait_type not in inscription_traits:
                        traits_usage[trait_type]["none"]["count"] += 1

                num_traits = len(inscription_traits)
                trait_count_distribution[num_traits] = trait_count_distribution.get(num_traits, 0) + 1
                inscription_collection.append(formatted_traits)
                break

    traits_usage_statistics = {}
    for trait_type, traits in traits_usage.items():
        traits_usage_statistics[trait_type] = {}

        trait_type_count = sum(usage_info["count"] for usage_info in traits.values())

        for trait, usage_info in traits.items():
            trait_info = {
                "usage": f"{usage_info['count']} ({(usage_info['count'] / trait_type_count) * 100:.2f}%)",
                "rarity input": usage_info["rarity"]
            }

            if trait == 'none':
                trait_info["none_status"] = "Used" if usage_info['count'] > 0 else "Not used"

            traits_usage_statistics[trait_type][trait] = trait_info

    sorted_trait_count_distribution = dict(sorted(trait_count_distribution.items()))

    return inscription_collection, traits_usage_statistics, trait_count_distribution

# ======================== MAIN EXECUTION ========================
def main():
    print("Collection Metadata Generator\n")

    trait_types = sorted(os.listdir(ROOT_DIRECTORY), key=lambda x: (int(x.split('.')[0]), x.split('.')[1]))
    traits_dict = {}

    for trait_type in trait_types:
        trait_type_name = trait_type.split('. ')[1]
        traits = os.listdir(os.path.join(ROOT_DIRECTORY, trait_type))
        traits_dict[trait_type_name] = [trait.split('.')[0] for trait in traits]

    create_spreadsheet(traits_dict)

    input("\nPress Enter after you have updated the 'traits_info.xlsx' file with the required information...")

    all_traits_info, trait_number_mapping = load_traits_info()

    num_inscriptions = get_positive_integer("\nEnter the number of Inscriptions you want to generate metadata for: ")

    try:
        inscription_collection, traits_usage_statistics, trait_count_distribution = generate_inscriptions(all_traits_info, trait_number_mapping, num_inscriptions)

        inconsistencies = validate_inscription_avoidance(inscription_collection, all_traits_info, trait_number_mapping)

        if inconsistencies:
            print("\nInconsistencies found in trait avoidance rules:")
            for inconsistency in inconsistencies:
                print(f"- {inconsistency}")

        with open('metadata.json', 'w') as file:
            json.dump(inscription_collection, file, indent=4)

        traits_inscription_mapping = {}
        for trait_type, traits in all_traits_info.items():
            traits_inscription_mapping[trait_type] = {trait: info["inscription_id"] for trait, info in traits.items() if trait != 'none'}  # Exclude 'none' traits

        with open('traits.json', 'w') as file:
            json.dump(traits_inscription_mapping, file, indent=4)

        # Prepare the summary for trait_usage_statistics.json
        summary = {
            "Total_inscriptions": num_inscriptions,
            "Trait_count_distribution": {f"{count}_traits": f"{amount} inscriptions" for count, amount in trait_count_distribution.items()},
            "Traits_usage": traits_usage_statistics
        }

        with open('trait_usage_statistics.json', 'w') as file:
            json.dump(summary, file, indent=4)

        print("\nInscription generation is complete. Check 'metadata.json', 'traits.json', and 'trait_usage_statistics.json' for the collection.")
    except Exception as e:
        print(f"An error occurred during Inscription generation: {e}")

if __name__ == '__main__':
    main()
